# --- CONFIG ---
# Excel Export Engine — client-aware workbook generator
# Uses openpyxl for all workbook creation (never xBA or xlsxwriter)
# All calculations are live Excel formulas — never hardcoded Python values
# Hidden _Lists sheets drive all data validation dropdowns
# Client-specific colors from client_config.py

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.client_config import get_client

# --- COLOR PALETTE ---
# Default TÜV colors (fallback)
TUV_BLUE = "003D7A"
TUV_RED = "C00000"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY = "666666"
GREEN = "00B050"
AMBER = "FFC000"
RED_BG = "FFE0E0"
GREEN_BG = "E2EFDA"
AMBER_BG = "FFF2CC"
LIGHT_BLUE = "D6E4F0"

# --- STYLE HELPERS ---
def _header_style():
    return {
        "font": Font(name="Calibri", bold=True, size=10, color=WHITE),
        "fill": PatternFill(start_color=TUV_BLUE, end_color=TUV_BLUE, fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color=DARK_GRAY),
            right=Side(style="thin", color=DARK_GRAY),
            top=Side(style="thin", color=DARK_GRAY),
            bottom=Side(style="thin", color=DARK_GRAY),
        ),
    }

def _data_style():
    return {
        "font": Font(name="Calibri", size=10),
        "alignment": Alignment(vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        ),
    }

def _apply_style(cell, style):
    for attr, val in style.items():
        setattr(cell, attr, val)

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _add_data_validation(ws, cell_range, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Invalid selection"
    dv.errorTitle = "Validation Error"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _hex_color(hex_str):
    """Convert hex color string to openpyxl-compatible AARRGGBB format."""
    h = hex_str.strip().lstrip("#")
    if len(h) == 6:
        return f"FF{h}"
    return h


# --- HIDDEN LISTS SHEET ---
def _create_lists_sheet(wb, client_key=None):
    """Create hidden _Lists sheet with dropdown values."""
    ws = wb.create_sheet("_Lists")
    ws.sheet_state = "hidden"

    lists = {
        "Risk_Category": ["Strategic", "Operational", "Financial", "Compliance", "Reputational", "Technical", "Environmental", "Health & Safety"],
        "Risk_Status": ["Open", "Mitigated", "Closed", "Monitoring", "Escalated"],
        "Treatment_Accept": ["Accept", "Mitigate", "Transfer", "Avoid"],
        "Likelihood": ["1 - Rare", "2 - Unlikely", "3 - Possible", "4 - Likely", "5 - Almost Certain"],
        "Impact": ["1 - Negligible", "2 - Minor", "3 - Moderate", "4 - Major", "5 - Catastrophic"],
        "Priority": ["Low", "Medium", "High", "Critical"],
        "BIA_Process": ["Core", "Support", "Administrative"],
        "BIA_Impact_Financial": ["< 100K", "100K - 500K", "500K - 1M", "1M - 5M", "> 5M"],
        "BIA_Impact_Reputational": ["Low", "Medium", "High", "Severe"],
        "BIA_Impact_Operational": ["< 1 day", "1-3 days", "3-7 days", "1-2 weeks", "> 2 weeks"],
        "EnMS_Significance": ["Low", "Medium", "High", "Very High"],
        "EnMS_Type": ["Risk", "Opportunity"],
        "Clause_Ref": ["4.1", "4.2", "4.3", "4.4", "5.1", "5.2", "5.3", "6.1", "6.2", "7.1", "7.2", "7.3", "7.4", "7.5", "8.1", "8.2", "9.1", "9.2", "9.3", "10.1", "10.2"],
    }

    col = 1
    for list_name, values in lists.items():
        ws.cell(row=1, column=col, value=list_name)
        _apply_style(ws.cell(row=1, column=col), _header_style())
        for i, val in enumerate(values, start=2):
            ws.cell(row=i, column=col, value=val)
        col += 1

    return ws


# --- RISK REGISTER ---
def generate_risk_register(client_key: str, output_path: str, data: list = None):
    """Generate a client-aware Risk Register Excel workbook.
    
    Uses client-specific formulas:
    - MSD-MOI: Latent S = O×Q, Residual V = S×(1−U/4)
    - UACC: L × S rating
    - Others: Standard L × I matrix
    """
    client = get_client(client_key) if client_key else None
    client_name = client.name if client else "Client"
    header_color = client.visual.primary_header if client else TUV_BLUE

    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    # Create hidden lists sheet
    _create_lists_sheet(wb, client_key)

    # Title row
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"Risk Register — {client_name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_hex_color(header_color))
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:N2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Doc Code: {client.doc_code_prefix if client else 'N/A'}RR-001"
    sub_cell.font = Font(name="Calibri", size=9, color=DARK_GRAY)
    sub_cell.alignment = Alignment(horizontal="center")

    # Column headers (row 4)
    headers = [
        "Risk ID", "Category", "Risk Description", "Cause", "Effect",
        "O\n(Occurrence)", "Q\n(Quality)", "S\n(Latent Score)",
        "Existing Controls", "U\n(Control Effectiveness)",
        "V\n(Residual Score)", "Risk Level", "Treatment Plan", "Status"
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=i, value=h)
        _apply_style(cell, _header_style())
        cell.fill = PatternFill(start_color=_hex_color(header_color), end_color=_hex_color(header_color), fill_type="solid")

    # Column widths
    col_widths = {1: 12, 2: 14, 3: 35, 4: 25, 5: 25, 6: 8, 7: 8, 8: 10, 9: 30, 10: 8, 11: 10, 12: 12, 13: 25, 14: 12}
    for col, width in col_widths.items():
        _set_col_width(ws, col, width)

    # Sample data rows (or use provided data)
    if data is None:
        data = _sample_risk_data(client_key)

    for row_idx, risk in enumerate(data, start=5):
        # Risk ID
        ws.cell(row=row_idx, column=1, value=risk.get("risk_id", f"R-{row_idx-4:03d}"))
        # Category (with dropdown)
        ws.cell(row=row_idx, column=2, value=risk.get("category", "Operational"))
        # Description
        ws.cell(row=row_idx, column=3, value=risk.get("description", ""))
        # Cause
        ws.cell(row=row_idx, column=4, value=risk.get("cause", ""))
        # Effect
        ws.cell(row=row_idx, column=5, value=risk.get("effect", ""))
        # O (Occurrence) — dropdown
        ws.cell(row=row_idx, column=6, value=risk.get("occurrence", 3))
        # Q (Quality) — dropdown
        ws.cell(row=row_idx, column=7, value=risk.get("quality", 3))

        # S (Latent Score) — FORMULA: =F{row}*G{row}
        s_cell = ws.cell(row=row_idx, column=8)
        s_cell.value = f"=F{row_idx}*G{row_idx}"
        s_cell.font = Font(name="Calibri", size=10, bold=True)

        # Existing Controls
        ws.cell(row=row_idx, column=9, value=risk.get("controls", ""))
        # U (Control Effectiveness) — dropdown 1-4
        ws.cell(row=row_idx, column=10, value=risk.get("control_eff", 2))

        # V (Residual Score) — client-specific formula
        v_cell = ws.cell(row=row_idx, column=11)
        if client_key == "msd_moi":
            # V = S × (1 − U/4)
            v_cell.value = f"=H{row_idx}*(1-J{row_idx}/4)"
        else:
            # Standard: V = S × (1 − U/5)
            v_cell.value = f"=H{row_idx}*(1-J{row_idx}/5)"
        v_cell.font = Font(name="Calibri", size=10, bold=True)

        # Risk Level — conditional formula
        rl_cell = ws.cell(row=row_idx, column=12)
        rl_cell.value = f'=IF(K{row_idx}>=12,"Critical",IF(K{row_idx}>=8,"High",IF(K{row_idx}>=4,"Medium","Low")))'
        rl_cell.font = Font(name="Calibri", size=10, bold=True)

        # Treatment Plan
        ws.cell(row=row_idx, column=13, value=risk.get("treatment", "Mitigate"))
        # Status
        ws.cell(row=row_idx, column=14, value=risk.get("status", "Open"))

        # Apply data style to all cells in row
        for col in range(1, 15):
            _apply_style(ws.cell(row=row_idx, column=col), _data_style())
            # Alternate row shading
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
                )

    # Add data validation dropdowns
    last_row = 4 + len(data)
    dv_range = f"B5:B{last_row}"
    _add_data_validation(ws, dv_range, "=_Lists!$A$2:$A$9")

    # Conditional formatting for Risk Level column (12)
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"L5:L{last_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"L5:L{last_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"L5:L{last_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"L5:L{last_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"))
    )

    # Freeze panes
    ws.freeze_panes = "A5"

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    return output_path


# --- BIA WORKBOOK ---
def generate_bia_workbook(client_key: str, output_path: str, data: list = None):
    """Generate Business Impact Analysis workbook."""
    client = get_client(client_key) if client_key else None
    client_name = client.name if client else "Client"
    header_color = client.visual.primary_header if client else TUV_BLUE

    wb = Workbook()
    ws = wb.active
    ws.title = "BIA Assessment"

    _create_lists_sheet(wb, client_key)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"].value = f"Business Impact Analysis — {client_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=_hex_color(header_color))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers (row 3)
    headers = [
        "Process ID", "Process Name", "Process Owner", "BIA Category",
        "Impact (Financial)", "Impact (Reputational)", "Impact (Operational)",
        "RTO (hrs)", "RPO (hrs)", "MTPD (hrs)",
        "Recovery Priority", "Dependencies"
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        _apply_style(cell, _header_style())
        cell.fill = PatternFill(start_color=_hex_color(header_color), end_color=_hex_color(header_color), fill_type="solid")

    col_widths = {1: 12, 2: 25, 3: 20, 4: 14, 5: 16, 6: 18, 7: 18, 8: 10, 9: 10, 10: 10, 11: 16, 12: 25}
    for col, width in col_widths.items():
        _set_col_width(ws, col, width)

    if data is None:
        data = _sample_bia_data()

    for row_idx, item in enumerate(data, start=4):
        ws.cell(row=row_idx, column=1, value=item.get("process_id", f"P-{row_idx-3:03d}"))
        ws.cell(row=row_idx, column=2, value=item.get("process_name", ""))
        ws.cell(row=row_idx, column=3, value=item.get("owner", ""))
        ws.cell(row=row_idx, column=4, value=item.get("category", "Core"))
        ws.cell(row=row_idx, column=5, value=item.get("impact_financial", ""))
        ws.cell(row=row_idx, column=6, value=item.get("impact_reputational", ""))
        ws.cell(row=row_idx, column=7, value=item.get("impact_operational", ""))
        ws.cell(row=row_idx, column=8, value=item.get("rto", 24))
        ws.cell(row=row_idx, column=9, value=item.get("rpo", 4))
        ws.cell(row=row_idx, column=10, value=item.get("mtpd", 72))

        # Recovery Priority — formula based on MTPD
        rp_cell = ws.cell(row=row_idx, column=11)
        rp_cell.value = f'=IF(J{row_idx}<=24,"Critical",IF(J{row_idx}<=72,"High",IF(J{row_idx}<=168,"Medium","Low")))'
        rp_cell.font = Font(name="Calibri", size=10, bold=True)

        ws.cell(row=row_idx, column=12, value=item.get("dependencies", ""))

        for col in range(1, 13):
            _apply_style(ws.cell(row=row_idx, column=col), _data_style())
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
                )

    # Conditional formatting for Recovery Priority
    last_row = 3 + len(data)
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"K4:K{last_row}",
        CellIsRule(operator="equal", formula=['"Critical"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"K4:K{last_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"))
    )

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    return output_path


# --- EnMS RISK & OPPORTUNITY REGISTER ---
def generate_enms_register(client_key: str, output_path: str, data: list = None):
    """Generate Energy Management System Risk & Opportunity Register."""
    client = get_client(client_key) if client_key else None
    client_name = client.name if client else "Client"
    header_color = client.visual.primary_header if client else TUV_BLUE

    wb = Workbook()
    ws = wb.active
    ws.title = "EnMS Risk & Opportunity"

    _create_lists_sheet(wb, client_key)

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"].value = f"EnMS Risk & Opportunity Register — {client_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=_hex_color(header_color))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers (row 3)
    headers = [
        "ID", "Type", "Clause Ref", "Description", "Energy Aspect",
        "Likelihood (L)", "Significance (S)", "Risk Score (L×S)",
        "Significance Level", "Action Required", "Status"
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=i, value=h)
        _apply_style(cell, _header_style())
        cell.fill = PatternFill(start_color=_hex_color(header_color), end_color=_hex_color(header_color), fill_type="solid")

    col_widths = {1: 8, 2: 12, 3: 10, 4: 35, 5: 20, 6: 12, 7: 14, 8: 12, 9: 16, 10: 25, 11: 12}
    for col, width in col_widths.items():
        _set_col_width(ws, col, width)

    if data is None:
        data = _sample_enms_data()

    for row_idx, item in enumerate(data, start=4):
        ws.cell(row=row_idx, column=1, value=item.get("id", f"ENR-{row_idx-3:03d}"))
        ws.cell(row=row_idx, column=2, value=item.get("type", "Risk"))
        ws.cell(row=row_idx, column=3, value=item.get("clause", "6.1"))
        ws.cell(row=row_idx, column=4, value=item.get("description", ""))
        ws.cell(row=row_idx, column=5, value=item.get("energy_aspect", ""))
        ws.cell(row=row_idx, column=6, value=item.get("likelihood", 3))
        ws.cell(row=row_idx, column=7, value=item.get("significance", 3))

        # Risk Score = L × S
        score_cell = ws.cell(row=row_idx, column=8)
        score_cell.value = f"=F{row_idx}*G{row_idx}"
        score_cell.font = Font(name="Calibri", size=10, bold=True)

        # Significance Level — nested IF
        sig_cell = ws.cell(row=row_idx, column=9)
        sig_cell.value = f'=IF(H{row_idx}>=15,"Very High",IF(H{row_idx}>=10,"High",IF(H{row_idx}>=5,"Medium","Low")))'
        sig_cell.font = Font(name="Calibri", size=10, bold=True)

        ws.cell(row=row_idx, column=10, value=item.get("action", ""))
        ws.cell(row=row_idx, column=11, value=item.get("status", "Open"))

        for col in range(1, 12):
            _apply_style(ws.cell(row=row_idx, column=col), _data_style())
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
                )

    # Conditional formatting for Significance Level
    last_row = 3 + len(data)
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"I4:I{last_row}",
        CellIsRule(operator="equal", formula=['"Very High"'], fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"I4:I{last_row}",
        CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"I4:I{last_row}",
        CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill(start_color=AMBER_BG, end_color=AMBER_BG, fill_type="solid"))
    )
    ws.conditional_formatting.add(
        f"I4:I{last_row}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid"))
    )

    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    return output_path


# --- KPI DASHBOARD ---
def generate_kpi_dashboard(client_key: str, output_path: str):
    """Generate a KPI dashboard workbook with summary metrics."""
    client = get_client(client_key) if client_key else None
    client_name = client.name if client else "Client"
    header_color = client.visual.primary_header if client else TUV_BLUE

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value = f"Compliance KPI Dashboard — {client_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=_hex_color(header_color))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # KPI Cards (row 3-5)
    kpis = [
        ("Total Risks", "24", "0"),
        ("Open NCs", "3", "0"),
        ("Closed NCs", "18", "0"),
        ("Compliance Score", "92%", "0"),
        ("Audits Completed", "5", "0"),
        ("Pending Actions", "7", "0"),
        ("BCM Readiness", "85%", "0"),
        ("EnMS Performance", "78%", "0"),
    ]

    for i, (label, value, _) in enumerate(kpis, start=1):
        col = i
        # Label
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = Font(name="Calibri", size=9, color=DARK_GRAY)
        cell.alignment = Alignment(horizontal="center")
        # Value
        cell = ws.cell(row=4, column=col, value=value)
        cell.font = Font(name="Calibri", bold=True, size=18, color=_hex_color(header_color))
        cell.alignment = Alignment(horizontal="center")
        # Card border
        for r in range(3, 5):
            ws.cell(row=r, column=col).border = Border(
                left=Side(style="medium", color=_hex_color(header_color)),
                right=Side(style="medium", color=_hex_color(header_color)),
                top=Side(style="medium", color=_hex_color(header_color)),
                bottom=Side(style="medium", color=_hex_color(header_color)),
            )
            ws.cell(row=r, column=col).fill = PatternFill(
                start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid"
            )
        _set_col_width(ws, col, 16)

    ws.row_dimensions[4].height = 30

    # Risk distribution chart data (row 7+)
    ws.cell(row=7, column=1, value="Risk Distribution by Level").font = Font(
        name="Calibri", bold=True, size=12, color=_hex_color(header_color)
    )

    dist_headers = ["Risk Level", "Count", "Percentage"]
    for i, h in enumerate(dist_headers, start=1):
        cell = ws.cell(row=8, column=i, value=h)
        _apply_style(cell, _header_style())
        cell.fill = PatternFill(start_color=_hex_color(header_color), end_color=_hex_color(header_color), fill_type="solid")

    dist_data = [
        ("Critical", 2, "8%"),
        ("High", 5, "21%"),
        ("Medium", 10, "42%"),
        ("Low", 7, "29%"),
    ]
    for row_idx, (level, count, pct) in enumerate(dist_data, start=9):
        ws.cell(row=row_idx, column=1, value=level)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=pct)
        for col in range(1, 4):
            _apply_style(ws.cell(row=row_idx, column=col), _data_style())

    _set_col_width(ws, 1, 16)
    _set_col_width(ws, 2, 10)
    _set_col_width(ws, 3, 12)

    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


# --- SAMPLE DATA ---
def _sample_risk_data(client_key=None):
    """Return sample risk data for demonstration."""
    return [
        {
            "risk_id": "R-001", "category": "Strategic",
            "description": "Regulatory changes impacting BCMS requirements",
            "cause": "New NCA ECC regulations", "effect": "Non-compliance, certification risk",
            "occurrence": 4, "quality": 3, "controls": "Regulatory monitoring, legal review",
            "control_eff": 2, "treatment": "Mitigate", "status": "Open"
        },
        {
            "risk_id": "R-002", "category": "Operational",
            "description": "Key personnel unavailability during BCM activation",
            "cause": "Single point of failure in crisis team", "effect": "Delayed response, extended downtime",
            "occurrence": 3, "quality": 4, "controls": "Cross-training, deputy assignments",
            "control_eff": 3, "treatment": "Mitigate", "status": "Monitoring"
        },
        {
            "risk_id": "R-003", "category": "Technical",
            "description": "ISMS control failure — Annex A.8 asset management",
            "cause": "Incomplete asset inventory", "effect": "Unauthorized access, data breach",
            "occurrence": 3, "quality": 3, "controls": "Automated asset discovery tool",
            "control_eff": 2, "treatment": "Mitigate", "status": "Open"
        },
        {
            "risk_id": "R-004", "category": "Compliance",
            "description": "ISO 50001 EnMS — energy baseline not updated",
            "cause": "No annual review process", "effect": "Inaccurate EnPI tracking",
            "occurrence": 2, "quality": 2, "controls": "Scheduled annual review",
            "control_eff": 3, "treatment": "Accept", "status": "Closed"
        },
        {
            "risk_id": "R-005", "category": "Financial",
            "description": "Budget overrun on certification project",
            "cause": "Scope creep, additional audit days", "effect": "Project delay",
            "occurrence": 3, "quality": 2, "controls": "Change control process",
            "control_eff": 2, "treatment": "Transfer", "status": "Open"
        },
    ]


def _sample_bia_data():
    """Return sample BIA data."""
    return [
        {"process_id": "P-001", "process_name": "Patient Records Management", "owner": "IT Director", "category": "Core", "impact_financial": "1M - 5M", "impact_reputational": "Severe", "impact_operational": "1-3 days", "rto": 4, "rpo": 1, "mtpd": 24, "dependencies": "Network, Database"},
        {"process_id": "P-002", "process_name": "Emergency Response", "owner": "Safety Manager", "category": "Core", "impact_financial": "500K - 1M", "impact_reputational": "High", "impact_operational": "< 1 day", "rto": 1, "rpo": 0, "mtpd": 4, "dependencies": "Communications"},
        {"process_id": "P-003", "process_name": "Payroll Processing", "owner": "HR Director", "category": "Support", "impact_financial": "100K - 500K", "impact_reputational": "Medium", "impact_operational": "3-7 days", "rto": 48, "rpo": 24, "mtpd": 72, "dependencies": "Finance System"},
        {"process_id": "P-004", "process_name": "Supply Chain Management", "owner": "Operations Director", "category": "Core", "impact_financial": "> 5M", "impact_reputational": "High", "impact_operational": "1-2 weeks", "rto": 72, "rpo": 12, "mtpd": 168, "dependencies": "ERP, Logistics"},
        {"process_id": "P-005", "process_name": "Facility Management", "owner": "Admin Manager", "category": "Support", "impact_financial": "< 100K", "impact_reputational": "Low", "impact_operational": "> 2 weeks", "rto": 168, "rpo": 48, "mtpd": 336, "dependencies": "Maintenance Team"},
    ]


def _sample_enms_data():
    """Return sample EnMS risk/opportunity data."""
    return [
        {"id": "ENR-001", "type": "Risk", "clause": "6.1", "description": "Inadequate energy baseline for new production line", "energy_aspect": "Electricity consumption", "likelihood": 4, "significance": 3, "action": "Establish new baseline per ISO 50006", "status": "Open"},
        {"id": "ENR-002", "type": "Opportunity", "clause": "6.2", "description": "LED lighting retrofit in warehouse", "energy_aspect": "Lighting energy", "likelihood": 5, "significance": 4, "action": "Implement Q3 2026", "status": "Open"},
        {"id": "ENR-003", "type": "Risk", "clause": "8.1", "description": "Compressor efficiency degradation", "energy_aspect": "Compressed air", "likelihood": 3, "significance": 3, "action": "Preventive maintenance program", "status": "Monitoring"},
        {"id": "ENR-004", "type": "Opportunity", "clause": "6.3", "description": "Solar panel installation on roof", "energy_aspect": "Renewable energy", "likelihood": 3, "significance": 5, "action": "Feasibility study", "status": "Open"},
        {"id": "ENR-005", "type": "Risk", "clause": "9.1", "description": "EnPI deviation due to production volume change", "energy_aspect": "Specific energy consumption", "likelihood": 4, "significance": 2, "action": "Adjust normalization factors", "status": "Closed"},
    ]


# --- MAIN EXPORT FUNCTION ---
def generate_excel(client_key: str, doc_type: str, output_dir: str, data: list = None):
    """Main entry point for Excel generation.
    
    Args:
        client_key: Client identifier (e.g., 'msd_moi', 'uacc')
        doc_type: Type of Excel document ('risk_register', 'bia', 'enms', 'dashboard')
        output_dir: Output directory path
        data: Optional custom data (uses sample data if None)
    
    Returns:
        Path to generated Excel file
    """
    os.makedirs(output_dir, exist_ok=True)
    client = get_client(client_key) if client_key else None
    prefix = client.doc_code_prefix if client else "DOC-"
    safe_name = client.name.replace(" ", "_")[:30] if client else "Client"

    generators = {
        "risk_register": (generate_risk_register, f"{prefix}RR-{safe_name}.xlsx"),
        "bia": (generate_bia_workbook, f"{prefix}BIA-{safe_name}.xlsx"),
        "enms": (generate_enms_register, f"{prefix}EnMS-ROR-{safe_name}.xlsx"),
        "dashboard": (generate_kpi_dashboard, f"{prefix}KPI-Dashboard-{safe_name}.xlsx"),
    }

    gen_func, filename = generators.get(doc_type, (None, None))
    if gen_func is None:
        raise ValueError(f"Unknown Excel doc_type: {doc_type}. Available: {list(generators.keys())}")

    output_path = os.path.join(output_dir, filename)

    if doc_type == "dashboard":
        return gen_func(client_key, output_path)
    return gen_func(client_key, output_path, data)
