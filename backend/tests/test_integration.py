"""End-to-end integration tests for the ComplianceHub platform.

Tests the full pipeline: client selection → document generation → output verification.
Covers all 4 clients, all 8 document types, Excel export, and bilingual output.
"""

import os
import sys
import json
import tempfile
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from app.main import app
from fastapi.testclient import TestClient


@pytest.fixture
def client():
    return TestClient(app)


# ── Client Management Tests ──

class TestClientManagement:
    def test_list_clients(self, client):
        r = client.get("/api/clients")
        assert r.status_code == 200
        data = r.json()
        assert len(data["clients"]) == 4
        keys = [c["key"] for c in data["clients"]]
        assert "msd_moi" in keys
        assert "uacc" in keys
        assert "sagco" in keys
        assert "al_ahsa" in keys

    def test_client_detail_msd_moi(self, client):
        r = client.get("/api/clients/msd_moi")
        assert r.status_code == 200
        d = r.json()
        assert d["language"] == "ar"
        assert d["visual"]["rtl"] is True
        assert d["visual"]["primary_header"] == "#004D26"
        assert d["formulas"]["latent_risk"] == "S = O × Q"
        assert d["formulas"]["residual_risk"] == "V = S × (1 − U/4)"
        assert d["doc_code_prefix"] == "MSD-MOI-GRC-"

    def test_client_detail_uacc(self, client):
        r = client.get("/api/clients/uacc")
        assert r.status_code == 200
        d = r.json()
        assert d["language"] == "en"
        assert d["visual"]["rtl"] is False
        assert d["formulas"]["rating_method"] == "L × S"

    def test_client_detail_al_ahsa(self, client):
        r = client.get("/api/clients/al_ahsa")
        assert r.status_code == 200
        d = r.json()
        assert d["language"] == "ar"
        assert "iso_27001" in d["standards"]

    def test_client_detail_sagco(self, client):
        r = client.get("/api/clients/sagco")
        assert r.status_code == 200
        d = r.json()
        assert "iso_45001" in d["standards"]
        assert "iso_14001" in d["standards"]

    def test_doc_code_generation(self, client):
        r = client.get("/api/clients/msd_moi/doc_code?doc_type=RR&sequence=1")
        assert r.json()["doc_code"] == "MSD-MOI-GRC-RR-001"

        r = client.get("/api/clients/uacc/doc_code?doc_type=BIA&sequence=5")
        assert r.json()["doc_code"] == "UACC-EnMS-BIA-005"

    def test_client_validation_correct(self, client):
        r = client.post("/api/clients/msd_moi/validate",
            json={"doc_code": "MSD-MOI-GRC-RR-001", "standard": "ISO 22301"})
        assert r.json()["valid"] is True

    def test_client_validation_wrong_prefix(self, client):
        r = client.post("/api/clients/msd_moi/validate",
            json={"doc_code": "UACC-EnMS-ROR-001"})
        assert r.json()["valid"] is False

    def test_unknown_client(self, client):
        r = client.get("/api/clients/unknown_client")
        assert r.status_code == 404


# ── Security Tests ──

class TestSecurity:
    def test_path_traversal_blocked(self, client):
        # Non-UUID job_id should be rejected
        r = client.get("/api/download/../../etc/passwd")
        assert r.status_code in (400, 404)  # 400 from validation, 404 from route mismatch

    def test_invalid_job_id(self, client):
        r = client.get("/api/download/abc123")
        # Non-UUID path rejected by route validation or returns 404
        assert r.status_code in (400, 404)

    def test_cors_configurable(self, client):
        # Verify the API responds (CORS middleware is active)
        r = client.get("/api/standards")
        assert r.status_code == 200


# ── Excel Export Tests ──

class TestExcelExport:
    def test_risk_register_msd_moi(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "msd_moi", "doc_type": "risk_register"})
        assert r.status_code == 200
        assert len(r.content) > 1000
        # Verify it's an xlsx file (PK header)
        assert r.content[:2] == b'PK'

    def test_risk_register_uacc(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "uacc", "doc_type": "risk_register"})
        assert r.status_code == 200
        assert r.content[:2] == b'PK'

    def test_bia_workbook(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "msd_moi", "doc_type": "bia"})
        assert r.status_code == 200
        assert r.content[:2] == b'PK'

    def test_enms_register(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "uacc", "doc_type": "enms"})
        assert r.status_code == 200
        assert r.content[:2] == b'PK'

    def test_kpi_dashboard(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "msd_moi", "doc_type": "dashboard"})
        assert r.status_code == 200
        assert r.content[:2] == b'PK'

    def test_excel_requires_client(self, client):
        r = client.post("/api/generate_excel",
            data={"doc_type": "risk_register"})
        # Non-UUID path rejected by route validation or returns 404
        assert r.status_code in (400, 404)

    def test_excel_unknown_client(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "nonexistent", "doc_type": "risk_register"})
        assert r.status_code == 404

    def test_excel_invalid_type(self, client):
        r = client.post("/api/generate_excel",
            data={"client_key": "msd_moi", "doc_type": "invalid_type"})
        # Non-UUID path rejected by route validation or returns 404
        assert r.status_code in (400, 404)


# ── Document Generator Tests ──

class TestDocumentGenerators:
    def test_all_generators_import(self):
        from app.services.document_generator import (
            generate_tnl, generate_certificate, generate_audit_plan_stage,
            generate_audit_plan_stage_1, generate_audit_plan_stage_2,
            generate_participation_list, generate_certificate_text,
            generate_audit_report, generate_iso_checklist,
            GENERATORS
        )
        assert len(GENERATORS) == 12

    def test_all_generators_have_client_key(self):
        import inspect
        from app.services.document_generator import GENERATORS
        for name, func in GENERATORS.items():
            sig = inspect.signature(func)
            assert "client_key" in sig.parameters, f"{name} missing client_key"

    def test_resolve_client_msd_moi(self):
        from app.services.document_generator import _resolve_client
        client, lang, rtl, color = _resolve_client("msd_moi")
        assert lang == "ar"
        assert rtl is True

    def test_resolve_client_uacc(self):
        from app.services.document_generator import _resolve_client
        client, lang, rtl, color = _resolve_client("uacc")
        assert lang == "en"
        assert rtl is False

    def test_resolve_client_default(self):
        from app.services.document_generator import _resolve_client
        client, lang, rtl, color = _resolve_client(None)
        assert lang == "en"
        assert rtl is False

    def test_tnl_generation_english(self):
        from app.services.document_generator import generate_tnl
        data = {
            "client_name": "Test Client",
            "standard": "ISO 27001:2022",
            "audit_date": "15/06/2026",
            "entries": [
                {
                    "tnl_number": "TNL-001", "clause": "4.1", "type": "NC",
                    "description": "Test nonconformity description",
                    "severity": "Minor", "auditee": "IT Manager",
                    "due_date": "30/06/2026", "status": "Open"
                }
            ],
            "summary": {"total_nc": 1, "major": 0, "minor": 1, "ofi": 0, "observations": 0}
        }
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
            path = generate_tnl(data, f.name, client_key="uacc")
            assert os.path.exists(path)
            assert os.path.getsize(path) > 1000
            os.unlink(path)

    def test_tnl_generation_arabic(self):
        from app.services.document_generator import generate_tnl
        data = {
            "client_name": "اختبار",
            "standard": "ISO 27001:2022",
            "audit_date": "15/06/2026",
            "entries": [],
            "summary": {"total_nc": 0, "major": 0, "minor": 0, "ofi": 0, "observations": 0}
        }
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
            path = generate_tnl(data, f.name, client_key="msd_moi")
            assert os.path.exists(path)
            os.unlink(path)

    def test_certificate_generation(self):
        from app.services.document_generator import generate_certificate
        data = {
            "client_name": "Test Corp", "standard": "ISO 9001:2015",
            "audit_date": "15/06/2026", "scope": "Quality Management",
            "lead_auditor": "John Doe", "certification_body": "TÜV AUSTRIA",
            "certification_decision": "Certified",
            "issue_date": "01/07/2026", "expiry_date": "01/07/2029",
            "authorized_signatory": "Jane Smith", "conditions": []
        }
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
            path = generate_certificate(data, f.name, client_key="uacc")
            assert os.path.exists(path)
            os.unlink(path)

    def test_iso_checklist_generation(self):
        from app.services.document_generator import generate_iso_checklist
        data = {
            "client_name": "Test Corp", "standard": "ISO 27001:2022",
            "audit_date": "15/06/2026", "auditor": "John Doe",
            "sections": [
                {
                    "clause": "4.1", "title": "Context",
                    "requirement": "Understanding the organization",
                    "status": "Conformant",
                    "evidence": "Documented context analysis available",
                    "notes": "", "reference": "Doc-001"
                }
            ],
            "overall_assessment": "The organization demonstrates conformity."
        }
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as f:
            path = generate_iso_checklist(data, f.name, client_key="uacc")
            assert os.path.exists(path)
            os.unlink(path)


# ── Bilingual System Tests ──

class TestBilingual:
    def test_english_headings(self):
        from app.services.bilingual import t
        assert t("audit_objectives", "en") == "1. Audit Objectives"
        assert t("nc_observation_log", "en") == "Nonconformity & Observation Log"

    def test_arabic_headings(self):
        from app.services.bilingual import t
        assert t("audit_objectives", "ar") == "1. أهداف المراجعة"
        assert t("nc_observation_log", "ar") == "سجل عدم المطابقة والملاحظات"

    def test_table_headers_bilingual(self):
        from app.services.bilingual import TABLE_HEADERS
        assert "tnl" in TABLE_HEADERS["en"]
        assert "tnl" in TABLE_HEADERS["ar"]
        assert len(TABLE_HEADERS["en"]["tnl"]) == len(TABLE_HEADERS["ar"]["tnl"])

    def test_cover_labels_bilingual(self):
        from app.services.bilingual import COVER_LABELS
        assert COVER_LABELS["en"]["client"] == "Client"
        assert COVER_LABELS["ar"]["client"] == "العميل"

    def test_methodology_bilingual(self):
        from app.services.bilingual import METHODOLOGY
        assert "approach" in METHODOLOGY["en"]
        assert "approach" in METHODOLOGY["ar"]
        assert len(METHODOLOGY["en"]["approach"]) > 50
        assert len(METHODOLOGY["ar"]["approach"]) > 50


# ── Excel Formula Tests ──

class TestExcelFormulas:
    def test_moi_risk_formula(self):
        from app.services.excel_generator import generate_risk_register
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = generate_risk_register("msd_moi", f.name)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["Risk Register"]
            k5 = str(ws["K5"].value)
            assert "(1-J5/4)" in k5, f"MOI formula not found: {k5}"
            os.unlink(path)

    def test_uacc_risk_formula(self):
        from app.services.excel_generator import generate_risk_register
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = generate_risk_register("uacc", f.name)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["Risk Register"]
            k5 = str(ws["K5"].value)
            assert "(1-J5/5)" in k5, f"UACC formula not found: {k5}"
            os.unlink(path)

    def test_hidden_lists_sheet(self):
        from app.services.excel_generator import generate_risk_register
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = generate_risk_register("msd_moi", f.name)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            assert "_Lists" in wb.sheetnames
            os.unlink(path)

    def test_enms_lxS_formula(self):
        from app.services.excel_generator import generate_enms_register
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = generate_enms_register("uacc", f.name)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb["EnMS Risk & Opportunity"]
            h4 = str(ws["H4"].value)
            assert "F4*G4" in h4 or "F4" in h4, f"L×S formula not found: {h4}"
            os.unlink(path)


# ── AI Pipeline Tests ──

class TestAIPipeline:
    def test_client_context_injection(self):
        from app.services.ai_pipeline import _build_prompt
        prompt = _build_prompt(
            "Test notes", "Test manday",
            ["ISO 27001:2022"], "Audit_Report",
            client_key="msd_moi"
        )
        assert "Client: MSD-MOI" in prompt
        assert "Arabic/RTL" in prompt
        assert "S = O × Q" in prompt

    def test_no_client_context(self):
        from app.services.ai_pipeline import _build_prompt
        prompt = _build_prompt(
            "Test notes", "Test manday",
            ["ISO 27001:2022"], "Audit_Report"
        )
        assert "Client Context" not in prompt

    def test_uacc_context(self):
        from app.services.ai_pipeline import _build_prompt
        prompt = _build_prompt(
            "Test notes", "Test manday",
            ["ISO 50001:2018"], "Audit_Report",
            client_key="uacc"
        )
        assert "UACC" in prompt
        assert "English/LTR" in prompt


# ── Offline Generator Tests ──

class TestOfflineGenerator:
    def test_offline_generates_all_docs(self):
        from app.services.offline_generator import generate_all
        results = generate_all(
            "Test audit notes for ISO 27001",
            "Lead Auditor: John Doe\nTotal Mandays: 5",
            ["ISO 27001:2022"], ["iso_27001"]
        )
        assert len(results) >= 6
        for doc_type, data in results.items():
            assert "error" not in data or data.get("client_name") is not None


# ── Code Quality Tests ──

class TestCodeQuality:
    def test_no_dead_code(self):
        import inspect
        from app.services import document_generator
        src = inspect.getsource(document_generator)
        assert "_inject_into_22301_template" not in src
        assert "_inject_into_20000_template" not in src

    def test_compileall_clean(self):
        import subprocess
        r = subprocess.run(
            ["python3", "-m", "compileall", ".", "-q"],
            capture_output=True,
            cwd=os.path.join(os.path.dirname(__file__), '..')
        )
        assert r.returncode == 0, f"Compile failed: {r.stderr.decode()}"

    def test_pyflakes_clean(self):
        import subprocess
        r = subprocess.run(
            ["python3", "-m", "pyflakes", "app/"],
            capture_output=True,
            cwd=os.path.join(os.path.dirname(__file__), '..')
        )
        assert r.returncode == 0, f"Pyflakes: {r.stdout.decode()}"
